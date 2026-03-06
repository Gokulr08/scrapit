from pathlib import Path
from scraper.config import OUTPUT_DIR

SHEET_NAME = "data"


def _get_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        return Workbook, load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel output.\n"
            "Install with: pip install openpyxl"
        )


def save(data: dict, name: str, *, output_dir: str | None = None) -> str:
    base = Path(output_dir) if output_dir else OUTPUT_DIR
    base.mkdir(parents=True, exist_ok=True)
    out_file = base / f"{name}.xlsx"

    Workbook, load_workbook = _get_openpyxl()

    row_values = {k: str(v) for k, v in data.items()}
    keys = list(row_values.keys())

    if not out_file.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        for col, key in enumerate(keys, 1):
            ws.cell(row=1, column=col, value=key)
        for col, key in enumerate(keys, 1):
            ws.cell(row=2, column=col, value=row_values[key])
        wb.save(out_file)
        return str(out_file)

    wb = load_workbook(out_file)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        for col, key in enumerate(keys, start=1):
            ws.cell(row=1, column=col, value=key)
        for col, key in enumerate(keys, start=1):
            ws.cell(row=2, column=col, value=row_values[key])
        wb.save(out_file)
        return str(out_file)
    
    ws = wb[SHEET_NAME]
    max_col = ws.max_column
    existing_headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    existing_headers = [h for h in existing_headers if h is not None and str(h).strip()]

    new_keys = [k for k in keys if k not in existing_headers]
    all_headers = existing_headers + new_keys

    if new_keys:
        for col, key in enumerate(all_headers, start=1):
            ws.cell(row=1, column=col, value=key)

    next_row = ws.max_row + 1
    for col, key in enumerate(all_headers, start=1):
        value = row_values.get(key, "")
        ws.cell(row=next_row, column=col, value=value)

    wb.save(out_file)
    return str(out_file)